import os
import sys
from pathlib import Path
from docx import Document
from docx_replace_ms import docx_replace
import openpyxl
from datetime import datetime, date
import configparser


def load_config():
    """Загружает конфигурацию из INI файла с валидацией значений."""

    config = configparser.ConfigParser()
    config_file_path = os.path.join(os.getcwd(), 'WordGenFromExcel.ini')

    if not os.path.exists(config_file_path):
        print("INI файл не найден!")
        input("Нажмите Enter для выхода ...")
        exit(1)
    try:
        config.read(config_file_path, encoding='utf-8')
        template_name = config.get('PATHS', 'template_name')
        data_file_name = config.get('PATHS', 'data_file_name')

        # валидация
        if not template_name or not isinstance(template_name, str):
            raise ValueError("Название файла шаблона договора в ini файле указано не корректно")
        if not data_file_name or not isinstance(data_file_name, str):
            raise ValueError("Название файла эксел с данными в ini файле указано не корректно")
        if not template_name.lower().endswith('.docx'):
            raise ValueError("Название файла с шаблоном договора должен оканчиваться на .docx")
        if not data_file_name.lower().endswith('.xlsx'):
            raise ValueError("Название файла с данными экселе должен оканчиваться на .xlsx")
        if not os.path.exists(os.path.join(os.getcwd(), template_name)):
            raise ValueError(f"Не найден файл шаблона договора - {template_name}")
        if not os.path.exists(os.path.join(os.getcwd(), data_file_name)):
            raise ValueError(f"Не найден файл эксел с данными - {data_file_name}")

        return template_name, data_file_name

    except Exception as e:
        print(f"Ошибка: {e}.")
        print("Проверьте ini файл на корректное заполнение или убедитесь в наличии файла шаблона и файла с данными!")
        input("Нажмите Enter для выхода ...")
        exit(1)

def excel_to_dict(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    try:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Файл Excel не содержит данных")
        # Извлекаем заголовки из первой строки
        headers = []
        for cell in rows[0]:
            if cell and str(cell).strip():
                headers.append(str(cell).strip())
            else:
                break
        if not headers:
            raise ValueError("Все ячейки верхней строки Excel файла должны быть заполнены!")

        result = {}
        # Обработка строк данных (все строки кроме первой)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                item_name = row[0]
                attributes = {}
                for i in range(1, len(headers)):
                    value = row[i]
                    # Обработка дат
                    if isinstance(value, (datetime, date)):
                        attributes[headers[i]] = value.strftime("%d.%m.%Y")
                    elif value is None:
                        attributes[headers[i]] = ""
                    else:
                        attributes[headers[i]] = str(value).strip() or ""
                result[item_name] = attributes
        return result
    finally:
        wb.close()

def main():
    # Загрузка конфигурации
    template_name, data_file_name = load_config()
    # Конфигурация путей
    exe_dir = os.getcwd()

    template_path = os.path.join(exe_dir, template_name)
    xlsx_path = os.path.join(exe_dir, data_file_name)

    try:
        # Работа с Excel-данными
        replacements = excel_to_dict(xlsx_path)

        for doc_name, attributes in replacements.items():
            # Загрузка шаблона
            doc = Document(template_path)
            # Выполнение замен
            docx_replace(doc, **attributes)

            # Сохранение результата
            output_file = f"{doc_name}{Path(template_name).suffix}"
            output_path = os.path.join(exe_dir, output_file)
            doc.save(output_path)
            print(f"Создан документ: {output_path}")

    except Exception as e:
        print(f"КРИТИЧЕСКАЯ ОШИБКА: {str(e)}", file=sys.stderr)
        input("Нажмите Enter для выхода ...")
        sys.exit(1)

if __name__ == "__main__":
    main()
